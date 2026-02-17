"""Parse documents to extract text content."""
import io

def parse_document(file_bytes, filename):
    """Extract text from document."""
    ext = filename.lower().split('.')[-1]
    
    if ext == 'txt':
        return file_bytes.decode('utf-8')
    elif ext == 'pdf':
        return parse_pdf(file_bytes)
    elif ext == 'docx':
        return parse_docx(file_bytes)
    elif ext == 'xlsx':
        return parse_xlsx(file_bytes)
    else:
        raise ValueError(f"Unsupported format: {ext}")

def parse_pdf(file_bytes):
    """Extract text from PDF with OCR fallback."""
    import tempfile
    import os
    
    # Try pypdf first (fast)
    try:
        from pypdf import PdfReader
        pdf = PdfReader(io.BytesIO(file_bytes))
        text = []
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text.append(page_text)
        
        result = '\n'.join(text).strip()
        if result and len(result) > 50:  # If we got meaningful text
            return result
    except Exception as e:
        print(f"pypdf extraction failed: {e}")
    
    # Fallback to OCR if pypdf failed or returned insufficient text
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
        from PIL import Image, ImageEnhance, ImageFilter
        
        # Convert PDF to images at higher DPI for better chart/graph text extraction
        images = convert_from_bytes(file_bytes, dpi=300)
        text = []
        
        for page_num, page_image in enumerate(images, 1):
            try:
                # Preprocess image for better OCR on charts/graphs
                # Convert to grayscale
                gray = page_image.convert('L')
                # Enhance contrast
                enhancer = ImageEnhance.Contrast(gray)
                enhanced = enhancer.enhance(2.0)
                # Sharpen
                sharpened = enhanced.filter(ImageFilter.SHARPEN)
                
                # Try multiple OCR modes and combine results
                ocr_texts = []
                
                # Mode 6: Uniform text block (for paragraphs)
                try:
                    text_mode6 = pytesseract.image_to_string(
                        sharpened, 
                        lang='eng+heb+ara',
                        config='--psm 6 --oem 1'
                    )
                    if text_mode6.strip():
                        ocr_texts.append(text_mode6)
                except:
                    pass
                
                # Mode 11: Sparse text (for charts/graphs with scattered numbers)
                try:
                    text_mode11 = pytesseract.image_to_string(
                        sharpened, 
                        lang='eng+heb+ara',
                        config='--psm 11 --oem 1'
                    )
                    if text_mode11.strip():
                        ocr_texts.append(text_mode11)
                except:
                    pass
                
                # Combine unique text from both modes
                combined_text = '\n'.join(ocr_texts)
                if combined_text.strip():
                    text.append(f"[Page {page_num}]\n{combined_text}")
            except Exception as ocr_e:
                print(f"OCR failed for page {page_num}: {ocr_e}")
                continue
        
        result = '\n\n'.join(text).strip()
        if result:
            return result
        else:
            return "[PDF text extraction failed - document may be image-based or corrupted]"
            
    except Exception as e:
        print(f"OCR fallback failed: {e}")
        return "[PDF text extraction failed - OCR not available or document corrupted]"

def parse_docx(file_bytes):
    """Extract text from DOCX."""
    from docx import Document
    doc = Document(io.BytesIO(file_bytes))
    text = []
    for para in doc.paragraphs:
        text.append(para.text)
    return '\n'.join(text)

def parse_xlsx(file_bytes):
    """Extract text from XLSX."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes))
    text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            row_text = []
            for cell in row:
                if cell.value:
                    row_text.append(str(cell.value))
            if row_text:
                text.append(' | '.join(row_text))
    return '\n'.join(text)
